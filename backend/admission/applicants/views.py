import openpyxl
from io import BytesIO
from django.db.models import Count
from django.http import FileResponse
from django.conf import settings
from rest_framework.decorators import action, api_view, permission_classes, authentication_classes

from .models import Applicant, StudentExam
from .serializers import ApplicantWriteSerializer, ApplicantReadSerializer
from rest_framework import viewsets, status
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response
from datetime import date, time, datetime


class ApplicantViewSet(viewsets.ModelViewSet):
    queryset = Applicant.objects.all()

    def get_permissions(self):
        if self.action == "create":
            return [AllowAny()]
        return [IsAuthenticated()]

    def get_serializer_class(self):
        if self.action in ["create", "update", "partial_update"]:
            return ApplicantWriteSerializer
        return ApplicantReadSerializer

    def get_queryset(self):
        queryset = Applicant.objects.all()
        search: str = self.request.query_params.get('search', None)

        status_filters = self.request.query_params.get('status', [])
        enrollment_filters = self.request.query_params.get('enrollment', [])
        institute_filters = self.request.query_params.get('institute', [])
        sort_by = self.request.query_params.get('sort_by', None)
        order = self.request.query_params.get('order', None)

        if search is not None:
            if search.isdigit():
                queryset = queryset.filter(national_id__icontains=search)
            else:
                queryset = queryset.filter(arabic_name__icontains=search)

        if len(status_filters) > 0:
            normal_status_filters = status_filters.split(',')
            queryset = queryset.filter(status__in=normal_status_filters)

        if len(enrollment_filters) > 0:
            enrollment_filters = enrollment_filters.split(',')
            queryset = queryset.filter(enrollment__in=enrollment_filters)

        if len(institute_filters) > 0:
            normal_institute_filters = institute_filters.split(',')
            queryset = queryset.filter(institute__in=normal_institute_filters)

        if sort_by is not None:
            queryset = queryset.order_by(f"{order}{sort_by}")

        return queryset

    @action(detail=True, methods=["post"])
    def set_status(self, request, pk=None):
        new_status = request.data.get("status")
        if new_status in ["مرفوض", "مقبول"]:
            try:
                applicant = Applicant.objects.get(pk=pk)
                applicant.status = new_status
                applicant.save()
                return Response(data={"status": new_status}, status=status.HTTP_200_OK)
            except Applicant.DoesNotExist:
                return Response({"error": "متقدم غير موجود"}, status=status.HTTP_404_NOT_FOUND)
        return Response(status=status.HTTP_405_METHOD_NOT_ALLOWED)


@api_view(["GET"])
@permission_classes([AllowAny])
def get_application(request):
    national_id = request.GET.get("national_id", None)

    if national_id is not None:
        try:
            applicant = Applicant.objects.get(national_id=national_id)
            serializer = ApplicantReadSerializer(applicant, many=False, context={'request': request})
            return Response(serializer.data)
        except Applicant.DoesNotExist:
            return Response({'detail': 'موظف غير موجود'}, status=status.HTTP_404_NOT_FOUND)

    return Response({'detail': 'الرقم القومي مطلوب'}, status=status.HTTP_404_NOT_FOUND)


@api_view(["get"])
@permission_classes([IsAuthenticated])
def get_home_statistics(request):
    status_stats = list(Applicant.objects.values("status").annotate(count=Count("status")))
    status_stats.append({"status": "الكل", "count": Applicant.objects.count()})
    enrollment = Applicant.objects.values("enrollment").annotate(count=Count("enrollment"))
    return Response({"status_stats": status_stats, "enrollment": enrollment})


@api_view(["get"])
@permission_classes([AllowAny])
@authentication_classes([])
def get_student_exam(request):
    now = datetime.now().astimezone(settings.CAIRO_TZ)

    target_date = date(2025, 8, 31)
    start_time = time(18, 0)  # 6:00 PM
    end_time = time(19, 0)  # 7:00 PM

    if not (now.date() == target_date and start_time <= now.time() <= end_time):
        return Response(
            {"detail": "الاختبار غير متاح الآن. متاح فقط من 6:00 حتى 7:00 مساءً."},
            status=status.HTTP_423_LOCKED
        )

    national_id = request.GET.get("national_id")

    if not national_id:
        return Response(
            {"detail": "الرقم القومي مطلوب للتحقق من صلاحية الدخول إلى الاختبار."},
            status=status.HTTP_400_BAD_REQUEST
        )

    try:
        student = StudentExam.objects.get(national_id=national_id)
    except StudentExam.DoesNotExist:
        return Response(
            {"detail": "لم يتم العثور على طالب مسجل بهذا الرقم القومي."},
            status=status.HTTP_404_NOT_FOUND
        )

    if student.mark is not None:
        return Response(
            {"detail": "لا يمكن دخول الاختبار. هذا الطالب قد أنهى الاختبار مسبقاً."},
            status=status.HTTP_403_FORBIDDEN
        )

    try:
        name = Applicant.objects.get(national_id=national_id).arabic_name
    except Applicant.DoesNotExist:
        name = None

    return Response(
        {"name": name, "national_id": national_id},
        status=status.HTTP_200_OK
    )


@api_view(["post"])
@permission_classes([AllowAny])
@authentication_classes([])
def record_exam_mark(request):
    mark = request.data.get("mark", None)
    national_id = request.data.get("national_id", None)
    if mark is None or national_id is None:
        return Response(
            {"detail": "الرقم القومي والدرجة مطلوبين."},
            status=status.HTTP_400_BAD_REQUEST
        )

    student_exam = StudentExam.objects.get(national_id=national_id)
    if student_exam.mark is not None:
        return Response({"detail": "لقد تم تسجيل إجاباتك من قبل"}, status=status.HTTP_409_CONFLICT)

    student_exam.mark = mark
    student_exam.save()

    return Response(status=status.HTTP_200_OK)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def export_applicants_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Applicants"

    ws.sheet_view.rightToLeft = True

    columns = [
        ("الكود", "id"),
        ("الاسم باللغة العربية", "arabic_name"),
        ("الاسم باللغة الانجليزية", "english_name"),
        ("الديانة", "religion"),
        ("بلد الجنسية", "nationality"),
        ("محل الميلاد", "place_of_birth"),
        ("النوع", "gender"),
        ("المحافظة", "governorate"),
        ("المدينة", "city"),
        ("الرقم القومي", "national_id"),
        ("تاريخ الميلاد", "birthdate"),
        ("رقم الموبايل", "mobile"),
        ("رقم الموبايل (2)", "mobile2"),
        ("البريد الالكتروني", "email"),
        ("العنوان", "address"),
        ("المعهد", "institute"),
        ("اسم المعهد", "institute_name"),
        ("الشعبة", "division"),
        ("الالتحاق", "enrollment"),
        ("المجموع", "total_mark"),
        ("إجمالي الدرجات", "total_out_of"),
        ("نسبة الشهادة", "certificate_percentage"),
        ("سنة الحصول على الشهادة", "certificate_year"),
        ("التقدير", "grade"),
        ("الحالة", "status"),
        ("تاريخ التسجيل", "created_at"),
    ]

    # Header row
    for col_num, (header, _) in enumerate(columns, 1):
        ws.cell(row=1, column=col_num, value=header)

    queryset = Applicant.objects.all()

    search: str = request.query_params.get('search', None)

    status_filters = request.query_params.get('status', [])
    enrollment_filters = request.query_params.get('enrollment', [])
    institute_filters = request.query_params.get('institute', [])
    sort_by = request.query_params.get('sort_by', None)
    order = request.query_params.get('order', None)

    if search is not None:
        if search.isdigit():
            queryset = queryset.filter(national_id__icontains=search)
        else:
            queryset = queryset.filter(arabic_name__icontains=search)

    if len(status_filters) > 0:
        normal_status_filters = status_filters.split(',')
        queryset = queryset.filter(status__in=normal_status_filters)

    if len(enrollment_filters) > 0:
        enrollment_filters = enrollment_filters.split(',')
        queryset = queryset.filter(enrollment__in=enrollment_filters)

    if len(institute_filters) > 0:
        normal_institute_filters = institute_filters.split(',')
        queryset = queryset.filter(institute__in=normal_institute_filters)

    if sort_by is not None:
        queryset = queryset.order_by(f"{order}{sort_by}")

    # Write data
    for row_num, obj in enumerate(queryset, 2):
        for col_num, (_, field) in enumerate(columns, 1):
            value = getattr(obj, field, "")
            ws.cell(row=row_num, column=col_num, value=str(value))

    # Save to in-memory file
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return FileResponse(
        output,
        as_attachment=True,
        filename="applicants.xlsx",
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def export_records_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "درجات المتقدمين"

    ws.sheet_view.rightToLeft = True

    columns = [
        ("الاسم باللغة العربية", "arabic_name"),
        ("الرقم القومي", "national_id"),
        ("الشعبة", "division"),
        ("الدرجة", "mark"),
    ]

    # Header row
    for col_num, (header, _) in enumerate(columns, 1):
        ws.cell(row=1, column=col_num, value=header)

    queryset = StudentExam.objects.all()

    # Write data
    for row_num, obj in enumerate(queryset, 2):
        national_id = obj.national_id
        mark = obj.mark
        try:

            applicant = Applicant.objects.get(national_id=national_id)
            name = applicant.arabic_name
            division = applicant.division
        except Applicant.DoesNotExist:
            name = None
            division = None

        ws.cell(row=row_num, column=1, value=str(name) if name is not None else "")
        ws.cell(row=row_num, column=2, value=national_id)
        ws.cell(row=row_num, column=3, value=str(division) if division is not None else "")
        ws.cell(row=row_num, column=4, value=mark)

    # Save to in-memory file
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return FileResponse(
        output,
        as_attachment=True,
        filename="النتيجة.xlsx",
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
